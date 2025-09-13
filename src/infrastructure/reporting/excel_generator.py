import os
import io
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from src.domain.reporting.report_generator import IReportGenerator
from src.domain.entities.entities import Ticket

def safe_write(ws: Worksheet, cell_coord: str, value: Any):
    """
    Función auxiliar para escribir de forma segura en una celda,
    manejando celdas combinadas.
    """
    try:
        # Buscamos si la celda es parte de un rango combinado
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                # Es una celda combinada. Descombinamos, escribimos y volvemos a combinar.
                range_str = str(merged_range)
                ws.unmerge_cells(range_str)
                ws[cell_coord] = value
                ws.merge_cells(range_str)
                return
        
        # Si no está en un rango combinado, simplemente escribimos el valor.
        ws[cell_coord] = value
        
    except Exception as e:
        print(f"Error escribiendo en la celda {cell_coord}: {e}")
        # Si algo falla, es mejor dejar la celda como está que corromper el archivo.
        pass

class ExcelReportGenerator(IReportGenerator):
    def generar(self, datos: List[Any], header_data: Dict) -> bytes:
        tickets: List[Ticket] = datos
        
        try:
            wb = load_workbook("template_reporte.xlsx")
            ws = wb.active
        except FileNotFoundError:
            print("ERROR CRÍTICO: No se encontró el archivo 'template_reporte.xlsx'.")
            raise
            
        BASE_URL = os.getenv("BASE_URL", "")

        try:
            # --- 2. RELLENAMOS LA CABECERA USANDO LA FUNCIÓN SEGURA ---
            safe_write(ws, 'A4', header_data.get('generado_por', 'N/A'))
            
            if "fecha_rango" in header_data:
                fecha_texto = header_data["fecha_rango"]
            else:
                fecha_texto = f"{header_data.get('fecha_inicio').strftime('%d/%m/%Y')} - {header_data.get('fecha_fin').strftime('%d/%m/%Y')}"
                

            
            safe_write(ws, 'D4', fecha_texto) # Celda D4 para la fecha
            
            safe_write(ws, 'D6', len(tickets) if tickets else 0)

            # --- 3. RELLENAMOS LA TABLA DE DATOS ---
            start_row = 9
            if tickets:
                link_font = Font(color="0000FF", underline="single")
                
                for row_idx, ticket in enumerate(tickets, start_row):
                    ws.cell(row=row_idx, column=1, value=ticket.fecha.strftime('%d/%m/%Y'))
                    
                    cell = ws.cell(row=row_idx, column=2, value=f"#{ticket.numero_ticket}")
                    cell.hyperlink = f"{BASE_URL}{ticket.numero_ticket}"
                    cell.font = link_font
                    
                    ws.cell(row=row_idx, column=3, value=ticket.servicio)
                    ws.cell(row=row_idx, column=4, value=ticket.usuario_reporta)
                    ws.cell(row=row_idx, column=5, value=ticket.correo_usuario)
                    ws.cell(row=row_idx, column=6, value=ticket.empresa)

            # --- 4. GUARDAMOS EN MEMORIA ---
            virtual_workbook = io.BytesIO()
            wb.save(virtual_workbook)
            virtual_workbook.seek(0)
            return virtual_workbook.read()

        except Exception as e:
            print(f"!!! ERROR CRÍTICO AL RELLENAR EL EXCEL: {e}")
            raise

