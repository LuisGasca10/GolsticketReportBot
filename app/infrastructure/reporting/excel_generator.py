import os
import shutil
import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from app.domain.entities.entities import Ticket
from app.domain.reporting.report_generator import IReportGenerator

class ExcelReportGenerator(IReportGenerator):
    def generar(self, datos: List[Any], ruta_archivo: str, header_data: Dict) -> str:
        tickets: List[Ticket] = datos
        BASE_URL = os.getenv("BASE_URL", "")
        
        # --- 1. COPIAR LA PLANTILLA ---
        template_path = 'template_reporte.xlsx'
        output_path = ruta_archivo
        
        if not os.path.exists(template_path):
            raise FileNotFoundError("No se encontró el archivo 'template_reporte.xlsx' en la raíz del proyecto.")
            
        shutil.copy(template_path, output_path)

        # --- 2. ABRIR LA COPIA Y RELLENAR DATOS ---
        wb = load_workbook(output_path)
        ws = wb.active

        link_font = Font(color="0000FF", underline="single")

        # --- CORRECCIÓN ---
        # Formatear y mostrar el rango de fechas completo
        fecha_inicio = header_data.get('fecha_inicio', datetime.now())
        fecha_fin = header_data.get('fecha_fin', datetime.now())
        fecha_informe_str = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        
        ws['A4'] = header_data.get('generado_por', 'N/A')
        ws['D4'] = fecha_informe_str # Escribimos el rango de fechas

        if not tickets:
            ws['A9'] = "No se encontraron tickets para este periodo."
            wb.save(output_path)
            return output_path

        # --- 3. CALCULAR Y RELLENAR EL CONTADOR DE ACTIVIDADES ---
        df = pd.DataFrame([t.__dict__ for t in tickets])
        
        ws['D6'] = len(df) 

        # --- 4. RELLENAR LA TABLA DE TICKETS ---
        start_row = 9
        for row_idx, ticket in enumerate(tickets, start=start_row):
            # Mapeo a las columnas del Excel (A, B, C, D, E, F)
            ws.cell(row=row_idx, column=1, value=ticket.fecha.strftime('%d/%m/%Y')) # Columna A
            
            # Crear el hipervínculo en la columna del ticket (Columna B)
            ticket_cell = ws.cell(row=row_idx, column=2)
            ticket_cell.value = f'=HYPERLINK("{BASE_URL}{ticket.numero_ticket}", "{ticket.numero_ticket}")'
            ticket_cell.font = link_font
            
            ws.cell(row=row_idx, column=3, value=ticket.servicio) # Columna C
            ws.cell(row=row_idx, column=4, value=ticket.usuario_reporta) # Columna D
            ws.cell(row_idx, column=5, value=ticket.correo_usuario) # Columna E
            ws.cell(row_idx, column=6, value=ticket.empresa) # Columna F
        
        # Guardar los cambios en el archivo copiado
        wb.save(output_path)
        return output_path

